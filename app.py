import pandas as pd
import openpyxl
import streamlit as st
import io

st.set_page_config(page_title = "Client Returns", layout = "wide")
st.title("Upload the NAV Sheet")
         
def extract_sheet_data(file_path):
    wb = openpyxl.load_workbook(file_path,data_only = True)

    rows = []

    for name in wb.sheetnames:
        ws = wb[name]
        client = ws["B4"].value
        weighted_return = ws["B12"].value
        absolute_return = ws["B13"].value

        rows.append({
            "Client": client,
            "Weighted Return": weighted_return,
            "Absolute Return": absolute_return
        })
    df = pd.DataFrame(rows)

    return df

file = st.file_uploader("Upload Excel Sheet", type=["xlsx"])
if not file:
    st.info("Please upload a single Excel (.xlsx) file to begin.")
    st.stop()

with st.spinner("Reading and parsing workbook..."):
    df = extract_sheet_data(file)
st.dataframe(df, use_container_width=True)

output = io.BytesIO()

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Results")

output.seek(0)

st.download_button(
    label="⬇️ Download Results as Excel",
    data=output,
    file_name="client_returns.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
